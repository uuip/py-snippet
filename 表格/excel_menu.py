from openpyxl import Workbook
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# xl = load_workbook(r"C:\Users\sharp\Desktop\menu.xlsx")
# sheet =xl['Sheet1']
xl = Workbook()
sheet = xl.active
sheet.title = "s"

m = {"山东": {"济南", "青岛"}, "河北": {"石家庄", "邯郸"}}
resource = xl.create_sheet("resource")
start_cell = resource.active_cell  # 'A1'
row = coordinate_from_string(start_cell)[1]  # 1
row_ = row
column_letter = coordinate_from_string(start_cell)[0]  # A
column_index = column_index_from_string(column_letter)  # A-->1

for k, v in m.items():
    to_write = [k] + list(v)
    resource.append(to_write)
    start_column = get_column_letter(column_index + 1)
    end_column = get_column_letter(column_index + len(v))
    xl.defined_names.append(
        DefinedName(name=k, attr_text=f"resource!${start_column}${row}:${end_column}${row}")
    )
    row += 1

dv = DataValidation(type="list", formula1=f'"{",".join(m.keys())}"', allow_blank=True)
dv.add("A1:A1048576")
dv2 = DataValidation(type="list", formula1=f"=INDIRECT(${column_letter}{row_})", allow_blank=True)
dv2.add("B1:B1048576")
sheet.add_data_validation(dv)
sheet.add_data_validation(dv2)


xl.save(r"C:\Users\sharp\Desktop\menu2.xlsx")
